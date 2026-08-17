from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import requests
import psycopg2
import psycopg2.extras
import os
import json
import re
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = FastAPI(title="Refinery Contract Chatbot API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static files & root route
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("static/index.html")

# -- Config -------------------------------------------------------------------
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:password@host:5432/railway")
DINOIKI_API_KEY = os.getenv("DINOIKI_API_KEY", "")
DINOIKI_URL = "https://ai.dinoiki.com/v1/chat/completions"
AI_MODEL = "gpt-4o"

# -- Stopwords bahasa Indonesia ------------------------------------------------
STOPWORDS = {
    'apa', 'siapa', 'berapa', 'yang', 'dan', 'atau', 'dari', 'untuk',
    'dengan', 'adalah', 'ini', 'itu', 'ada', 'tidak', 'bisa', 'mau',
    'saya', 'kamu', 'dia', 'kami', 'kita', 'mereka', 'semua', 'sudah',
    'belum', 'sedang', 'akan', 'telah', 'pada', 'di', 'ke', 'oleh',
    'juga', 'hanya', 'lebih', 'paling', 'sangat', 'banyak', 'sedikit',
    'tampilkan', 'tunjukkan', 'cari', 'lihat', 'data', 'info', 'informasi',
    'list', 'daftar', 'total', 'jumlah', 'nilai', 'status', 'semua',
    'kontrak', 'vendor', 'tagihan', 'dokumen', 'progress', 'bulan', 'tahun'
}

# -- Helper: call dinoiki AI --------------------------------------------------
def call_ai(messages: list, max_tokens: int = 1500) -> str:
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {DINOIKI_API_KEY}"
    }
    payload = {
        "model": AI_MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": 0.3,
    }
    resp = requests.post(DINOIKI_URL, headers=headers, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    return data["choices"][0]["message"]["content"].strip()

# -- Dynamic Context Injection -------------------------------------------------
def smart_entity_search(user_message: str) -> str:
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        context = []
        found_ids = set()

        words = [w for w in re.findall(r'\b\w{3,}\b', user_message)
                 if w.lower() not in STOPWORDS]
        company_patterns = re.findall(r'\b(?:PT|CV|UD|TB|PD)\s+[\w\s]+', user_message, re.IGNORECASE)
        tag_patterns = re.findall(
            r'\b\d{2,3}-[A-Z]{1,3}-\d{3,}\b|\b\d{3}[A-Z]\d{3,}\b',
            user_message, re.IGNORECASE
        )
        search_terms = list(set(words + company_patterns + tag_patterns))

        for term in search_terms:
            term = term.strip()
            if len(term) < 3:
                continue

            cur.execute("""
                SELECT id_vendor, nama_vendor, status_vendor, score
                FROM vendor WHERE nama_vendor ILIKE %s LIMIT 3
            """, (f'%{term}%',))
            for v in cur.fetchall():
                key = f"vendor_{v[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[VENDOR] '{v[1]}' -> id_vendor={v[0]}, status={v[2]}, score={v[3]}")

            cur.execute("""
                SELECT k.id_kontrak, k.judul_kontrak, k.no_dokumen_kontrak,
                       k.direksi_pekerjaan, k.status_kontrak, k.tipe_kontrak, v.nama_vendor
                FROM kontrak k LEFT JOIN vendor v ON k.id_vendor = v.id_vendor
                WHERE k.judul_kontrak ILIKE %s OR k.no_dokumen_kontrak ILIKE %s
                   OR k.no_po_pr ILIKE %s OR k.direksi_pekerjaan ILIKE %s LIMIT 3
            """, (f'%{term}%', f'%{term}%', f'%{term}%', f'%{term}%'))
            for k in cur.fetchall():
                key = f"kontrak_{k[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[KONTRAK] '{k[1]}' -> id_kontrak={k[0]}, doc={k[2]}, direksi={k[3]}, status={k[4]}, tipe={k[5]}, vendor='{k[6]}'")

            cur.execute("""
                SELECT t.id_tagihan, t.nomor_tagihan, t.status_tagihan, t.nilai_tagihan, k.judul_kontrak
                FROM tagihan t LEFT JOIN kontrak k ON t.id_kontrak = k.id_kontrak
                WHERE t.nomor_tagihan ILIKE %s LIMIT 3
            """, (f'%{term}%',))
            for t in cur.fetchall():
                key = f"tagihan_{t[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[TAGIHAN] '{t[1]}' -> id_tagihan={t[0]}, status={t[2]}, nilai={t[3]}, kontrak='{t[4]}'")

            cur.execute("""
                SELECT p.id_padi, p.no_pembelian, p.judul_pembelian, p.nilai, v.nama_vendor
                FROM padi p LEFT JOIN vendor v ON p.id_vendor = v.id_vendor
                WHERE p.no_pembelian ILIKE %s OR p.judul_pembelian ILIKE %s LIMIT 3
            """, (f'%{term}%', f'%{term}%'))
            for p in cur.fetchall():
                key = f"padi_{p[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[PADI] '{p[2]}' -> id_padi={p[0]}, no_pembelian={p[1]}, nilai={p[3]}, vendor='{p[4]}'")

            cur.execute("""
                SELECT id_direksi_pekerjaan, nama, jabatan, sub_area
                FROM direksi_pekerjaan
                WHERE nama ILIKE %s OR jabatan ILIKE %s OR sub_area ILIKE %s
                LIMIT 3
            """, (f'%{term}%', f'%{term}%', f'%{term}%'))
            for d in cur.fetchall():
                key = f"direksi_{d[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(
                        f"[DIREKSI] '{d[1]}' -> id_direksi_pekerjaan={d[0]}, "
                        f"jabatan={d[2]}, sub_area={d[3]}"
                    )

            cur.execute("""
                SELECT dr.id_report, dr.tag_number, dr.deskripsi,
                       dr.tanggal_laporan, dr.disiplin, dr.status_pekerjaan
                FROM daily_report dr
                WHERE dr.tag_number ILIKE %s OR dr.deskripsi ILIKE %s
                ORDER BY dr.tanggal_laporan DESC
                LIMIT 3
            """, (f'%{term}%', f'%{term}%'))
            for r in cur.fetchall():
                key = f"report_{r[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(
                        f"[DAILY REPORT] tag={r[1]} -> id_report={r[0]}, "
                        f"tanggal={r[3]}, disiplin={r[4]}, "
                        f"status={r[5]}, deskripsi='{str(r[2])[:80]}'"
                    )

        conn.close()

        if context:
            result = "\n\nKONTEKS ENTITAS YANG DITEMUKAN DI DATABASE:\n"
            result += "(Gunakan informasi ini untuk memahami maksud user tanpa perlu klarifikasi)\n"
            result += "\n".join(context)
            return result

        return ""

    except Exception as e:
        return ""

# -- DB Schema context untuk AI -----------------------------------------------
SCHEMA_CONTEXT = """
Database PostgreSQL untuk sistem manajemen kontrak kilang minyak. Berikut skema tabel:

TABEL: profiles
Kolom: id, email, full_name, role (admin/pic/user/vendor/external), password_hash, created_at, updated_at, is_active, id_vendor

TABEL: vendor
Kolom: id_vendor, nama_vendor, npwp, alamat, pic_nama, pic_kontak, status_vendor (Active/Inactive/Blacklist), score, created_at, updated_at

TABEL: direksi_pekerjaan
Kolom: id_direksi_pekerjaan, nama, jabatan (Manager Maintenance Execution I/Manager Maintenance Execution II/Pengawas Pekerjaan), sub_area (SH Maintenance Area 5/SH Maintenance Area 6/SH Maintenance Area 7/Workshop), created_at, updated_at
Catatan: Tabel master nama & jabatan direksi/pengawas. Kolom direksi_pekerjaan di tabel kontrak berisi NAMA ORANG (string bebas), bukan kode area seperti MA5/MA6.

TABEL: program_kerja
Kolom: id_program_kerja, nama (Rutin/Non Rutin/TA/OH), created_at, updated_at

TABEL: planner
Kolom: id_planner, nama (P&S/OH/TA), created_at, updated_at

TABEL: kontrak
Kolom: id_kontrak, id_vendor, judul_kontrak, no_dokumen_kontrak, no_po_pr, no_irkap, direksi_pekerjaan,
  program_kerja, planner, kbo_bagian,
  tipe_kontrak (Lumpsum/Unit Price/TSA/LTSA/TSA/LTSA), status_kontrak (Pre-KOM/Aktif/Active/Selesai/Completed/Terminated),
  tanggal_spb_diterima, tanggal_terima_dokumen, tanggal_maksimal_kom, tanggal_mulai, tanggal_selesai,
  sla_kom_hari, estimasi_tanggal_kom, tanggal_kom, kom_terlambat, nilai_awal, durasi_kontrak_hari,
  progress_plan, progress_actual, aktivitas_saat_ini, kendala, disiplin, tkdn_percentage, tanggal_lkp,
  tanggal_mpl, tanggal_mpa, masa_pemeliharaan_hari,
  has_amendment, no_amandemen, tanggal_amandemen, jenis_amandemen, nilai_kontrak_baru, durasi_amandemen,
  tanggal_mulai_baru, tanggal_selesai_baru, alasan_perubahan,
  contract_documents (JSON), amendment_documents (JSON), s_curve_data (JSON),
  created_at, updated_at

TABEL: amandemen_kontrak
Kolom: id_amandemen, id_kontrak, nomor_urut, no_amandemen, tanggal_amandemen, jenis_amandemen,
  nilai_kontrak_baru, durasi_amandemen, tanggal_mulai_baru, tanggal_selesai_baru, alasan_perubahan,
  amendment_documents (JSON), created_at, updated_at

TABEL: tagihan
Kolom: id_tagihan, id_kontrak, nomor_tagihan, tanggal_tagihan, tipe_kontrak, termin, nilai_tagihan,
  status_tagihan, memo_required, tanggal_pengiriman_memo, dokumen_memo, dokumen_tagihan, catatan,
  created_at, updated_at

TABEL: sla_tagihan
Kolom: id, id_kontrak, id_tagihan,
  tgl_masuk_ba_joint_inspection, tgl_selesai_ba_joint_inspection,
  tgl_masuk_ba_commissioning, tgl_selesai_ba_commissioning,
  tgl_masuk_ba_penerimaan_material, tgl_selesai_ba_penerimaan_material,
  tgl_masuk_lkp, tgl_selesai_lkp,
  tgl_masuk_bast, tgl_selesai_bast,
  tgl_masuk_bakp, tgl_selesai_bakp,
  tgl_masuk_ivendor, tgl_selesai_ivendor,
  tgl_masuk_sa, tgl_selesai_sa,
  tgl_masuk_pa, tgl_selesai_pa,
  tgl_masuk_verifikasi, tgl_selesai_verifikasi,
  tgl_masuk_payment, tgl_selesai_payment,
  created_at, updated_at
Catatan: Tracking tanggal masuk & selesai per tahap untuk satu tagihan.

TABEL: sla_setting
Kolom: kode_tahap, batas_hari, warning_persen
Catatan: Konfigurasi batas hari dan ambang peringatan (%) per tahap SLA tagihan.

TABEL: progress_lumpsum
Kolom: id_progress, id_kontrak, milestone, persen, tanggal_update, evidence, created_at

TABEL: progress_unit_price
Kolom: id_progress, id_kontrak, nama_item, satuan, qty_rencana, qty_aktual, harga_satuan, tanggal_update, created_at

TABEL: monitoring_ltsa
Kolom: id_log, id_kontrak, tanggal_kunjungan, jenis_layanan (Preventive/Corrective/Standby),
  durasi_jam, sla_terpenuhi (Yes/No), keterangan, created_at

TABEL: padi
Kolom: id_padi, no_pembelian, tanggal, judul_pembelian, no_po_pr, nilai, id_vendor, link_pembelian,
  bagian, dokumen_pendukung, status_purchase (BAST), tanggal_bast, tanggal_sa_gr, tanggal_invoice,
  tanggal_payment_approval, tanggal_paid, catatan_status, created_at, updated_at

TABEL: dokumen_approval
Kolom: id_dokumen, id_kontrak, tipe_dokumen (Evident/Report/Persetujuan), nama_dokumen,
  deskripsi_dokumen, file_path, file_url, nama_file, tipe_file, ukuran_file,
  status_approval (Pending/Approved/Rejected), catatan_reviewer, uploaded_by, reviewed_by,
  reviewed_at, created_at, updated_at

TABEL: konfigurasi_sistem
Kolom: id_setting, nama_setting, nilai_setting, deskripsi, updated_at

TABEL: daily_report
Kolom: id_report, tanggal_laporan, disiplin, kategori, deskripsi, direksi, tag_number,
  status_pekerjaan, catatan, pengirim_wa, raw_text, created_at
Catatan: Laporan harian kegiatan maintenance.

Relasi penting:
- vendor.id_vendor -> kontrak.id_vendor (1 vendor banyak kontrak)
- kontrak.id_kontrak -> tagihan.id_kontrak
- kontrak.id_kontrak -> sla_tagihan.id_kontrak
- tagihan.id_tagihan -> sla_tagihan.id_tagihan
- kontrak.id_kontrak -> amandemen_kontrak.id_kontrak
- kontrak.id_kontrak -> progress_lumpsum.id_kontrak
- kontrak.id_kontrak -> progress_unit_price.id_kontrak
- kontrak.id_kontrak -> monitoring_ltsa.id_kontrak
- kontrak.id_kontrak -> dokumen_approval.id_kontrak
- vendor.id_vendor -> padi.id_vendor

NILAI ENUM & PILIHAN YANG VALID:

1. TIPE KONTRAK: 'Lumpsum', 'Unit Price', 'TSA', 'LTSA', 'TSA/LTSA'

2. STATUS KONTRAK: 'Pre-KOM', 'Aktif', 'Active', 'Selesai', 'Completed', 'Terminated'
   PENTING: 'Aktif' dan 'Active' adalah sinonim — gunakan OR saat filter status aktif.
   Begitu pula 'Selesai' dan 'Completed'. Contoh: WHERE status_kontrak IN ('Aktif', 'Active')

3. DISIPLIN: 'Instrument', 'Instrumentasi', 'Stationary', 'Electrical', 'Rotating', 'Alat Berat'
   PENTING: 'Instrument' dan 'Instrumentasi' adalah nilai yang sama — gunakan ILIKE '%instru%' atau OR.

4. DIREKSI PEKERJAAN: Field kontrak.direksi_pekerjaan berisi NAMA ORANG/JABATAN (bukan kode MA5/MA6).
   Untuk filter berdasarkan area, JOIN ke tabel direksi_pekerjaan dan filter kolom sub_area.
   Sub_area valid: 'SH Maintenance Area 5', 'SH Maintenance Area 6', 'SH Maintenance Area 7', 'Workshop'

5. JENIS AMANDEMEN: 'Nilai', 'Waktu', 'Nilai dan Waktu'
6. STATUS APPROVAL: 'Pending', 'Approved', 'Rejected'
7. STATUS VENDOR: 'Active', 'Inactive', 'Blacklist'
8. JENIS LAYANAN LTSA: 'Preventive', 'Corrective', 'Standby'

9. TAHAPAN SLA TAGIHAN — 11 tahap (kolom di tabel sla_tagihan, format: tgl_masuk_X / tgl_selesai_X):
   1. ba_joint_inspection     2. ba_commissioning        3. ba_penerimaan_material
   4. lkp                     5. bast                    6. bakp
   7. ivendor                 8. sa                      9. pa
   10. verifikasi             11. payment
   Tahap dianggap selesai jika tgl_selesai_X tidak NULL. Durasi = tgl_selesai - tgl_masuk.

10. KATEGORI DAILY REPORT: 'Corrective Maintenance', 'Preventive Maintenance', 'Plant Patrol',
    'Progress', 'Challenge Session', 'Support'

11. STATUS PEKERJAAN (daily_report): 'Done', 'In Progress', 'Waiting Material', 'Pending', '-'

12. PROGRAM KERJA: 'Rutin', 'Non Rutin', 'TA', 'OH'

13. PLANNER: 'P&S', 'OH', 'TA'

14. STATUS PURCHASE PADI: 'BAST'
"""

BASE_SYSTEM_PROMPT = (
    "Kamu adalah asisten cerdas untuk sistem manajemen kontrak kilang minyak.\n"
    "Kamu dapat menjawab pertanyaan bisnis dalam bahasa Indonesia secara natural "
    "dan mengkonversinya ke query SQL PostgreSQL.\n\n"
    + SCHEMA_CONTEXT +
    "\nATURAN KETAT:\n"
    "1. HANYA boleh generate query SELECT, TIDAK boleh UPDATE, DELETE, INSERT, DROP, ALTER, TRUNCATE, dll\n"
    "2. TIDAK boleh query SELECT * (tanpa kolom spesifik) - selalu tentukan kolom yang relevan\n"
    "3. Selalu gunakan LIMIT maksimal 1000 baris\n"
    "4. Gunakan JOIN yang tepat antar tabel\n"
    "5. Format angka nilai kontrak dalam format Indonesia (Rp)\n"
    "\n⚠️ ATURAN KRITIS — WAJIB QUERY DATABASE, JANGAN JAWAB DARI INGATAN:\n"
    "SETIAP pertanyaan yang menyebut data spesifik (tanggal, nilai, status, progress, nama, jumlah)\n"
    "WAJIB menggunakan type='query' dan SQL — TIDAK BOLEH dijawab dari ingatan atau estimasi.\n"
    "Ini berlaku meskipun konteks entitas sudah diinjeksi. Konteks entitas HANYA boleh dipakai\n"
    "untuk mendapatkan ID (id_kontrak, id_vendor, dll) yang dimasukkan ke WHERE clause SQL.\n"
    "JANGAN PERNAH jadikan nilai dari konteks entitas (status, tanggal, nilai) sebagai jawaban final.\n"
    "type='narrative' HANYA boleh dipakai untuk:\n"
    "- Sapaan, Ucapan terima kasih, Pertanyaan tentang kemampuan AI\n"
    "Semua pertanyaan lain → type='query'.\n"
    "\nATURAN INTERPRETASI ENTITAS:\n"
    '- Jika ada blok "KONTEKS ENTITAS YANG DITEMUKAN DI DATABASE" -> gunakan ID-nya untuk WHERE clause SQL\n'
    "- Jika user menyebut nama yang diawali PT/CV/UD -> cari di vendor.nama_vendor\n"
    "- Jika user menyebut kode seperti MA5, KOM-001 -> cari di direksi_pekerjaan atau no_dokumen_kontrak\n"
    "- Jika entitas tidak ditemukan di konteks -> baru boleh minta klarifikasi\n"
    "\nFORMAT RESPONS JSON:\n"
    "Kamu HARUS selalu merespons dalam format JSON seperti ini:\n"
    "{\n"
    '  "type": "query" | "clarification" | "narrative" | "error",\n'
    '  "sql": "query SQL jika type=query",\n'
    '  "explanation": "penjelasan dalam bahasa Indonesia apa yang akan dilakukan query ini",\n'
    '  "narrative_hint": "bagaimana cara menarasikan hasilnya nanti",\n'
    '  "chart_suggestion": null | "bar" | "line" | "pie" | "doughnut",\n'
    '  "chart_config": null | {"x_column": "...", "y_column": "...", "label": "..."},\n'
    '  "clarification_question": "pertanyaan klarifikasi jika type=clarification",\n'
    '  "message": "pesan untuk user"\n'
    "}\n"
    "\nDETEKSI CHART:\n"
    '- Jika pertanyaan menyebut "grafik", "chart", "trend", "perbandingan", "distribusi", "per bulan/tahun" -> suggest chart\n'
    "- bar chart: perbandingan kategori\n"
    "- line chart: data time-series\n"
    "- pie/doughnut: distribusi persentase\n"
)

# -- Laporan System Prompt -----------------------------------------------------
LAPORAN_SYSTEM_PROMPT = (
    "Kamu adalah parser laporan harian maintenance kilang minyak.\n"
    "Tugasmu mengekstrak data dari teks laporan narasi ke dalam format JSON terstruktur.\n\n"
    "DISIPLIN YANG VALID: Electrical, Instrument, Rotating, Stationary, Alat Berat\n\n"
    "KATEGORI YANG VALID:\n"
    "- Corrective Maintenance\n"
    "- Preventive Maintenance\n"
    "- Plant Patrol\n"
    "- Progress\n"
    "- Challenge Session\n"
    "- Support\n\n"
    "STATUS YANG VALID: Done, In Progress, Waiting Material, Pending, -\n"
    "MAPPING STATUS (terapkan tepat seperti ini):\n"
    "  (done), DONE, Done, selesai → Done\n"
    "  (ip), (in progress), In Progress, in progress, sedang dikerjakan → In Progress\n"
    "  waiting material, wm → Waiting Material\n"
    "  pending → Pending\n"
    "  Jika tidak ada keterangan status → -\n\n"
    "DIREKSI (area kerja) — normalisasi ke format standar:\n"
    "  'Maintenance Area 7' / 'Area 7' / 'MA 7' / 'Bagian 7' → 'MA7'\n"
    "  'Maintenance Area 5' / 'Area 5' / 'MA 5' / 'Bagian 5' → 'MA5'\n"
    "  'Maintenance Area 6' / 'Area 6' / 'MA 6' / 'Bagian 6' → 'MA6'\n"
    "  'Workshop' → 'Workshop'\n"
    "Jika tidak ada informasi direksi, gunakan string kosong.\n\n"
    "TAG NUMBER: Kode identifikasi equipment/alat yang biasanya ada di awal deskripsi item,\n"
    "dipisah dengan titik dua (:) atau spasi. Contoh: 101-P-105, 104-P-107, 101A514, 105-FV-020.\n"
    "Format umum: [area]-[tipe]-[nomor] atau [area][kode][nomor].\n"
    "Jika tidak ada tag number, gunakan string kosong.\n\n"
    "ATURAN EKSTRAKSI:\n"
    "1. Satu item pekerjaan = satu entri JSON\n"
    "2. Deteksi tanggal dari teks laporan (format DD/MM/YYYY, DD Bulan YYYY, dsb)\n"
    "3. Deteksi disiplin dari header laporan\n"
    "4. Deteksi direksi dari header laporan, normalisasi ke MA5/MA6/MA7/Workshop sesuai aturan di atas\n"
    "5. Petakan setiap item ke kategori yang sesuai\n"
    "6. Ekstrak status menggunakan MAPPING STATUS di atas\n"
    "7. Ekstrak tag number dari awal deskripsi item jika ada\n"
    "8. Deskripsi diisi tanpa tag number (tag number sudah dipisah di field tag_number)\n"
    "9. Catatan: info tambahan yang relevan (target tanggal, detail teknis, dll)\n\n"
    "ATURAN MULTI-TAG (penting):\n"
    "Jika satu baris menyebut beberapa tag sekaligus, buat SATU entri per tag.\n"
    "Contoh: 'Perbaikan fireproofing: 105-P-506, 105-P-508, 105-P-507 (in progress)'\n"
    "→ 3 entri terpisah, masing-masing dengan tag berbeda, deskripsi & status sama.\n\n"
    "ATURAN ABAIKAN BARIS BERIKUT (jangan buat entri JSON):\n"
    "- Baris template/placeholder, contoh: 'Tag Number/Equipment/Func. Loc: Aktifitas (status)', '...', '..'\n"
    "- Baris sub-header equipment, contoh: '* Equipment : Transmitter', '* Equipment : Junction Box'\n"
    "- Baris kosong atau hanya berisi tanda baca\n\n"
    "RESPONSE FORMAT — kembalikan HANYA array JSON, tanpa teks lain:\n"
    '[\n  {\n    "tanggal_laporan": "2026-05-26",\n    "disiplin": "Instrument",\n'
    '    "direksi": "MA7",\n    "kategori": "Plant Patrol",\n    "tag_number": "105-FV-020",\n'
    '    "deskripsi": "Plant Patrol control valve",\n'
    '    "status_pekerjaan": "Done",\n    "catatan": ""\n  }\n]\n\n'
    "PENTING: Kembalikan HANYA array JSON yang valid. Jangan tambahkan penjelasan apapun."
)

# -- Laporan Functions ---------------------------------------------------------
PLACEHOLDER_PATTERNS = [
    r'^tag\s*number',
    r'^func.*loc',
    r'^\.*$',
    r'^aktifitas',
    r'^\s*\.\.\.*\s*$',
]

def is_junk_entry(item: dict) -> bool:
    deskripsi  = item.get("deskripsi",  "").strip().lower()
    tag_number = item.get("tag_number", "").strip().lower()
    combined   = f"{tag_number} {deskripsi}".strip()
    for pat in PLACEHOLDER_PATTERNS:
        if re.search(pat, combined, re.IGNORECASE):
            return True
    if not deskripsi and not tag_number:
        return True
    return False

def parse_laporan_with_ai(raw_text: str) -> list:
    try:
        response = call_ai([
            {"role": "system", "content": LAPORAN_SYSTEM_PROMPT},
            {"role": "user",   "content": f"Parse laporan berikut:\n\n{raw_text}"}
        ], max_tokens=2000)

        json_match = re.search(r'\[[\s\S]*\]', response)
        if not json_match:
            return []

        parsed = json.loads(json_match.group())
        if not isinstance(parsed, list):
            return []
        return [item for item in parsed if not is_junk_entry(item)]

    except Exception as e:
        print(f"[PARSE LAPORAN ERROR] {e}")
        return []

def insert_daily_report(items: list, pengirim: str, raw_text: str) -> tuple:
    """Return: (success_count, error_msg | None, dup_warnings_list, saved_items_list)"""
    if not items:
        return 0, "Tidak ada item yang bisa diparse", [], []
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur  = conn.cursor()
        success      = 0
        skipped      = 0
        dup_warnings = []
        saved_items  = []

        for item in items:
            if not item.get("tanggal_laporan") or not item.get("disiplin") or not item.get("deskripsi"):
                skipped += 1
                continue

            # Cek duplikat sebelum INSERT
            cur.execute("""
                SELECT id_report FROM daily_report
                WHERE tanggal_laporan = %s
                  AND disiplin        = %s
                  AND tag_number      = %s
                  AND deskripsi       = %s
                LIMIT 1
            """, (
                item.get("tanggal_laporan"),
                item.get("disiplin", "-"),
                item.get("tag_number", ""),
                item.get("deskripsi", "-"),
            ))
            if cur.fetchone():
                label = item.get("tag_number") or item.get("deskripsi", "?")[:30]
                dup_warnings.append(label)
                continue

            cur.execute("""
                INSERT INTO daily_report
                    (tanggal_laporan, disiplin, direksi, kategori, tag_number, deskripsi,
                     status_pekerjaan, catatan, pengirim_wa, raw_text)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                item.get("tanggal_laporan"),
                item.get("disiplin",         "-"),
                item.get("direksi",          ""),
                item.get("kategori",         "-"),
                item.get("tag_number",       ""),
                item.get("deskripsi",        "-"),
                item.get("status_pekerjaan", "-"),
                item.get("catatan",          ""),
                pengirim,
                raw_text,
            ))
            saved_items.append(item)
            success += 1

        conn.commit()
        conn.close()
        if skipped:
            print(f"[INSERT LAPORAN] {skipped} item dilewati (field wajib kosong)")
        return success, None, dup_warnings, saved_items

    except Exception as e:
        print(f"[INSERT LAPORAN ERROR] {e}")
        return 0, str(e), [], []

# -- Models --------------------------------------------------------------------
class ChatRequest(BaseModel):
    message: str
    history: list = []
    pengirim: str = "web"

class LaporanRequest(BaseModel):
    raw_text: str
    pengirim: str = "web"

class DownloadRequest(BaseModel):
    sql: str
    filename: str = "data_export"

# -- DB Connection -------------------------------------------------------------
def get_db_connection():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database connection failed: {str(e)}")

# -- SQL Validator -------------------------------------------------------------
def validate_sql(sql: str) -> tuple:
    sql_upper = sql.upper().strip()

    dangerous = ["UPDATE", "DELETE", "INSERT", "DROP", "ALTER", "TRUNCATE", "CREATE", "GRANT", "REVOKE"]
    for op in dangerous:
        if re.search(r'\b' + op + r'\b', sql_upper):
            return False, f"Operasi {op} tidak diizinkan."

    if re.search(r'SELECT\s+\*', sql_upper):
        return False, "Query SELECT * tidak diizinkan."

    if not re.search(r'\bSELECT\b', sql_upper):
        return False, "Hanya query SELECT yang diizinkan."

    if "LIMIT" not in sql_upper:
        sql = sql.rstrip(";") + " LIMIT 1000"

    return True, sql

# -- Execute Query -------------------------------------------------------------
def execute_query(sql: str) -> tuple:
    valid, result = validate_sql(sql)
    if not valid:
        raise HTTPException(status_code=400, detail=result)

    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(result)
            rows    = cur.fetchall()
            columns = [desc[0] for desc in cur.description] if cur.description else []
            data    = [dict(row) for row in rows]
            return data, columns
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Query error: {str(e)}")
    finally:
        conn.close()

# -- Generate Narrative --------------------------------------------------------
def generate_narrative(data: list, columns: list, original_question: str, narrative_hint: str) -> str:
    if not data:
        return "Tidak ditemukan data yang sesuai dengan pertanyaan Anda."

    if len(data) <= 5 and len(columns) <= 5:
        data_str = json.dumps(data, default=str, ensure_ascii=False)
        try:
            return call_ai([
                {
                    "role": "system",
                    "content": "Kamu adalah asisten laporan bisnis. Jawab hanya dalam bahasa Indonesia yang profesional dan natural."
                },
                {
                    "role": "user",
                    "content": (
                        f'Pertanyaan user: "{original_question}"\n'
                        f"Hint narasi: {narrative_hint}\n"
                        f"Data hasil query: {data_str}\n\n"
                        "Buatkan narasi singkat dalam bahasa Indonesia. Maksimal 3 kalimat."
                    )
                }
            ], max_tokens=400)
        except Exception:
            return ""

    return ""

# -- Chat Endpoint -------------------------------------------------------------
@app.post("/api/chat")
async def chat(req: ChatRequest):
    # ── Deteksi #laporan ──────────────────────────────────────────────────────
    LAPORAN_TRIGGERS = ["#laporan","#Laporan", "#report", "#lpr"]
    matched_laporan  = None
    for trigger in LAPORAN_TRIGGERS:
        if req.message.lower().startswith(trigger):
            matched_laporan = trigger
            break

    if matched_laporan:
        laporan_text = req.message[len(matched_laporan):].strip()
        if not laporan_text:
            return {
                "type": "laporan_info",
                "message": (
                    "📋 Format pengiriman laporan:\n\n"
                    "#laporan [isi laporan]\n\n"
                    "Contoh:\n"
                    "#laporan Pekerjaan Rotating MA7 26 Mei 2026\n"
                    "Corrective Maintenance\n"
                    "1. 101-P-103: Perbaikan koneksi SAF (done)"
                ),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        items  = parse_laporan_with_ai(laporan_text)
        if not items:
            return {
                "type": "laporan_error",
                "message": (
                    "⚠️ Gagal memparse laporan.\n\n"
                    "Pastikan ada: tanggal, disiplin, dan daftar pekerjaan."
                ),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        success_count, error, dup_warnings, saved_items = insert_daily_report(items, req.pengirim, laporan_text)
        if error:
            return {
                "type": "laporan_error",
                "message": f"⚠️ Gagal menyimpan laporan: {error}",
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        summary = {}
        for item in saved_items:
            key = f"{item.get('disiplin', '-')} - {item.get('kategori', '-')}"
            summary[key] = summary.get(key, 0) + 1
        summary_lines = "\n".join([f"  • {k}: {v} item" for k, v in summary.items()])
        dup_note = f"\n\n⚠️ {len(dup_warnings)} item duplikat dilewati." if dup_warnings else ""

        return {
            "type": "laporan_success",
            "message": (
                f"✅ Laporan berhasil disimpan!\n\n"
                f"📋 Total: {success_count} kegiatan tercatat\n\n"
                f"Rincian:\n{summary_lines}{dup_note}"
            ),
            "data": saved_items,
            "columns": ["tanggal_laporan", "disiplin", "direksi", "kategori", "tag_number", "deskripsi", "status_pekerjaan"],
            "chart": None,
            "narrative": None,
            "sql": None,
            "row_count": success_count
        }

    # ── Normal chat flow ──────────────────────────────────────────────────────
    dynamic_context = smart_entity_search(req.message)
    system_prompt   = BASE_SYSTEM_PROMPT + dynamic_context

    messages = [{"role": "system", "content": system_prompt}]
    for h in req.history[-10:]:
        messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": req.message})

    try:
        raw = call_ai(messages, max_tokens=1500)

        json_match = re.search(r'\{[\s\S]*\}', raw)
        if not json_match:
            return {
                "type": "narrative", "message": raw, "data": None,
                "columns": [], "chart": None, "narrative": raw,
                "sql": None, "row_count": 0
            }

        parsed        = json.loads(json_match.group())
        response_type = parsed.get("type", "narrative")

        if response_type == "clarification":
            return {
                "type": "clarification",
                "message": parsed.get("clarification_question", parsed.get("message", "")),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        if response_type == "error":
            return {
                "type": "error", "message": parsed.get("message", "Terjadi kesalahan."),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        if response_type == "query" and parsed.get("sql"):
            sql          = parsed["sql"]
            valid, val_result = validate_sql(sql)
            if not valid:
                return {
                    "type": "error", "message": val_result, "data": None,
                    "columns": [], "chart": None, "narrative": None,
                    "sql": None, "row_count": 0
                }

            data, columns = execute_query(val_result)
            row_count     = len(data)

            narrative = ""
            if 0 < row_count <= 5 and len(columns) <= 5:
                narrative = generate_narrative(data, columns, req.message, parsed.get("narrative_hint", ""))

            chart = None
            if parsed.get("chart_suggestion") and row_count > 0:
                chart = {"type": parsed["chart_suggestion"], "config": parsed.get("chart_config", {})}

            serializable_data = []
            for row in data:
                clean_row = {}
                for k, v in row.items():
                    clean_row[k] = v.isoformat() if hasattr(v, 'isoformat') else v
                serializable_data.append(clean_row)

            return {
                "type": "query", "message": parsed.get("explanation", ""),
                "data": serializable_data, "columns": columns, "chart": chart,
                "narrative": narrative, "sql": val_result, "row_count": row_count
            }

        narrative_msg = parsed.get("message", raw)
        DATA_KEYWORDS = {
            'tanggal', 'nilai', 'progress', 'status', 'berapa', 'kapan',
            'siapa', 'kontrak', 'tagihan', 'vendor', 'amandemen', 'selesai',
            'mulai', 'durasi', 'harga', 'bayar', 'laporan', 'sla'
        }
        question_words = set(re.findall(r'\b\w+\b', req.message.lower()))
        if question_words & DATA_KEYWORDS:
            narrative_msg += (
                "\n\n⚠️ *Catatan:* Jawaban ini belum diverifikasi dari database. "
                "Tanyakan lebih spesifik (misalnya sebut nama kontrak atau vendor) "
                "agar saya bisa mengambil data yang akurat."
            )
        return {
            "type": "narrative", "message": narrative_msg,
            "data": None, "columns": [], "chart": None,
            "narrative": narrative_msg, "sql": None, "row_count": 0
        }

    except json.JSONDecodeError:
        return {
            "type": "narrative", "message": raw, "data": None,
            "columns": [], "chart": None, "narrative": raw,
            "sql": None, "row_count": 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# -- Laporan Endpoint (dedicated) ----------------------------------------------
@app.post("/api/laporan")
async def submit_laporan(req: LaporanRequest):
    """Endpoint khusus untuk submit laporan harian dari web UI."""
    if not req.raw_text.strip():
        raise HTTPException(status_code=400, detail="Teks laporan tidak boleh kosong")

    items = parse_laporan_with_ai(req.raw_text)
    if not items:
        raise HTTPException(status_code=422, detail="Gagal memparse laporan. Pastikan ada tanggal, disiplin, dan daftar pekerjaan.")

    success_count, error, dup_warnings, saved_items = insert_daily_report(items, req.pengirim, req.raw_text)
    if error:
        raise HTTPException(status_code=500, detail=f"Gagal menyimpan: {error}")

    return {
        "success": True,
        "total_saved": success_count,
        "duplicates_skipped": len(dup_warnings),
        "items": saved_items
    }

# -- Download Excel ------------------------------------------------------------
@app.post("/api/download")
async def download_excel(req: DownloadRequest):
    data, columns = execute_query(req.sql)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Export"

    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(start_color="E8EEF8", end_color="E8EEF8", fill_type="solid")
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val  = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{req.filename}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/health")
def health():
    return {"status": "ok", "message": "Refinery Contract Chatbot API running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)